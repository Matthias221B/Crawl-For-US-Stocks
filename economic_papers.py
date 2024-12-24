import requests
from bs4 import BeautifulSoup
from datetime import datetime
from lxml import etree
from openpyxl import load_workbook

# 伪装为正常用户
headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36 Edg/131.0.0.0'}
# 基础url
basic_url = 'https://quotes.sina.com.cn/usstock/hq/income.php?s='
# 输入代码
us_paper = input('请输入想爬取的美股代码: ')
# 构建url
url = basic_url + us_paper
# 发送GET请求获取网页内容，并将响应内容存储在resq变量中
resq = requests.get(url, headers=headers)
# 设置响应内容的编码格式为utf-8，确保中文字符正常显示
resq.encoding = 'utf-8'
# 使用etree.HTML方法将HTML文本解析为一个HTML文档对象
e = etree.HTML(resq.content.decode('utf-8', errors='ignore'))



# 加载工作簿
wb = load_workbook(r'F:\us_equities\us_economic.xlsx')
# 选择要复制的工作表
source_ws = wb['Sheet1']
# 创建一个新的工作表，名称为us_paper
new_ws = wb.create_sheet(title=us_paper)

# 复制Sheet1的内容到新的工作表
for row in source_ws.iter_rows():
    for cell in row:
        new_ws[cell.coordinate].value = cell.value




# 创建一个包含10个空列表的列表
date_xpath = [[] for _ in range(10)]
# 使用XPath查找元素
basic_data = [elem.text.strip() for elem in e.xpath("//th[@class='black']/following-sibling::td")]
# 计算数据组数
group_data = len(basic_data) // 5
# 将数据打包
for i in range(group_data):
    for j in range(5):
        date_xpath[i].append(basic_data[i * 5 + j])




# 定义一个函数来添加字符"亿"
def add_yi(value):
    return f"{value}亿"
# 定义各项数据并在每个元素后面加上字符"亿"



total_revenue = list(map(add_yi, date_xpath[0]))
maori = list(map(add_yi, date_xpath[1]))
Earnings_before_interest_and_tax = list(map(add_yi, date_xpath[2]))
Net_profit_before_tax = list(map(add_yi, date_xpath[3]))
Consolidated_net_profit = list(map(add_yi, date_xpath[4]))
net_profit = list(map(add_yi, date_xpath[5]))
Net_profit_on_general_available_revenue = list(map(add_yi, date_xpath[6]))
basic_earnings_per_share = list(map(add_yi, date_xpath[7]))
diluted_earnings_per_share = list(map(add_yi, date_xpath[8]))
Earnings_before_interest_tax_depreciation_and_amortisation = list(map(add_yi, date_xpath[9]))



# 将数据写入Excel，从第二行B列开始
data = [
    total_revenue,
    maori,
    Earnings_before_interest_and_tax,
    Net_profit_before_tax,
    Consolidated_net_profit,
    net_profit,
    Net_profit_on_general_available_revenue,
    basic_earnings_per_share,
    diluted_earnings_per_share,
    Earnings_before_interest_tax_depreciation_and_amortisation
]



for row_idx, lst in enumerate(data, start=2):
    for col_idx, value in enumerate(lst, start=2):
        new_ws.cell(row=row_idx, column=col_idx, value=value)

# 保存工作簿
wb.save(r'F:\us_equities\us_economic.xlsx')